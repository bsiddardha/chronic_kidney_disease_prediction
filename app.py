from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, make_response, jsonify
import numpy as np
from io import StringIO, BytesIO
import csv
import pickle
import openpyxl
import os
import re
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
import seaborn as sns

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Change this to a secure secret key
model = pickle.load(open('Kidney.pkl', 'rb'))

HEALTHY_RANGES = {
    'sg': {'min': 1.005, 'max': 1.030, 'unit': 'g/cm³'},
    'hemo': {'min': 13.5, 'max': 17.5, 'unit': 'g/dL'},
    'rc': {'min': 4.5, 'max': 5.9, 'unit': 'million/µL'},
    'pc': {'min': 0, 'max': 5, 'unit': 'per HPF'},
    'al': {'min': 0, 'max': 0.3, 'unit': 'g/dL'},
}

def create_comparison_chart(user_values):
    """Create a comparison chart between user values and healthy ranges"""
    plt.figure(figsize=(10, 6))
    
    # Filter for numerical values that have reference ranges
    comparable_values = {k: v for k, v in user_values.items() if k in HEALTHY_RANGES}
    
    # Prepare data for plotting
    metrics = list(comparable_values.keys())
    user_vals = [comparable_values[m] for m in metrics]
    healthy_mins = [HEALTHY_RANGES[m]['min'] for m in metrics]
    healthy_maxs = [HEALTHY_RANGES[m]['max'] for m in metrics]
    
    # Create normalized values for better visualization
    normalized_user = []
    for i, metric in enumerate(metrics):
        min_val = HEALTHY_RANGES[metric]['min']
        max_val = HEALTHY_RANGES[metric]['max']
        range_size = max_val - min_val
        normalized_user.append((user_vals[i] - min_val) / range_size)
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot healthy range as a green band
    for i in range(len(metrics)):
        ax.fill_between([i-0.2, i+0.2], [0, 0], [1, 1], color='lightgreen', alpha=0.3)
    
    # Plot user values
    ax.scatter(range(len(metrics)), normalized_user, color='blue', s=100, zorder=5, label='Your Values')
    
    # Customize the plot
    plt.xticks(range(len(metrics)), [m.upper() for m in metrics], rotation=45)
    plt.ylabel('Normalized Values')
    plt.title('Your Values Compared to Healthy Ranges')
    
    # Add grid
    plt.grid(True, linestyle='--', alpha=0.7)
    
    # Save plot to BytesIO object
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=300)
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer
# Define the Excel file paths
user_data_file = 'user_data.xlsx'
prediction_data_file = 'prediction_data.xlsx'
user_login_data_file = 'user_login_data.xlsx'
user_signup_data_file = 'user_signup_data.xlsx'

# Initialize Excel files
def initialize_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(file_path)

initialize_excel_file(user_data_file, ['Username', 'Password'])
initialize_excel_file(prediction_data_file, ['Username', 'Timestamp', 'Specific Gravity', 'Hypertension', 'Hemoglobin', 'Diabetes Mellitus', 'Albumin', 'Appetite', 'Red Blood Cells', 'Pus Cell', 'Prediction'])
initialize_excel_file(user_login_data_file, ['Username', 'Timestamp'])
initialize_excel_file(user_signup_data_file, ['Username', 'Timestamp'])

# Add cache control headers to all responses
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

def username_exists(username):
    workbook = openpyxl.load_workbook(user_data_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return True
    return False

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        if 'loginUsername' in request.form and 'loginPassword' in request.form:
            username = request.form['loginUsername']
            password = request.form['loginPassword']
            
            workbook = openpyxl.load_workbook(user_data_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == username and check_password_hash(row[1], password):
                    log_user_login(username)
                    session['username'] = username
                    return redirect(url_for('index'))
            
            flash('Invalid username or password', 'error')
        else:
            flash('Please provide both username and password', 'error')
    
    return render_template('login_signup.html')

@app.route('/signup', methods=['POST'])
def signup():
    name = request.form['signupName']
    username = request.form['signupUsername']
    password = request.form['signupPassword']
    confirm_password = request.form['signupConfirmPassword']

    if not re.match(r'^[A-Za-z\s]+$', name):
        flash('Full name should only contain letters and spaces', 'error')
        return redirect(url_for('home'))
    
    
    if len(name) < 6:  # Added name length validation
        flash('Full name must be at least 6 characters long', 'error')
        return redirect(url_for('home'))

    if len(username) < 8:
        flash('Username must be at least 8 characters long', 'error')
        return redirect(url_for('home'))
    
    if not re.match(r'^[A-Za-z0-9]+$', username):
        flash('Username should only contain letters and numbers', 'error')
        return redirect(url_for('home'))

    if len(password) < 8:
        flash('Password must be at least 8 characters long', 'error')
        return redirect(url_for('home'))

    if password != confirm_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('home'))

    if username_exists(username):
        flash('Username already in use. Please choose a different username.', 'error')
        return redirect(url_for('home'))

    hashed_password = generate_password_hash(password)
    
    workbook = openpyxl.load_workbook(user_data_file)
    sheet = workbook.active
    sheet.append([username, hashed_password])
    workbook.save(user_data_file)

    log_user_signup(username)
    flash('Signup successful! Please log in.', 'success')
    return redirect(url_for('home'))

def log_user_login(username):
    workbook = openpyxl.load_workbook(user_login_data_file)
    sheet = workbook.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([username, timestamp])
    workbook.save(user_login_data_file)

def log_user_signup(username):
    workbook = openpyxl.load_workbook(user_signup_data_file)
    sheet = workbook.active
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([username, timestamp])
    workbook.save(user_signup_data_file)

@app.route('/index')
def index():
    if 'username' not in session:
        return redirect(url_for('home'))
    response = make_response(render_template('index.html'))
    return response

@app.route('/check_session')
def check_session():
    """Endpoint to check if session is still valid"""
    if 'username' in session:
        return jsonify({'valid': True})
    return jsonify({'valid': False})

@app.route('/contact')
def contact():
    if 'username' not in session:
        return redirect(url_for('home'))
    return render_template('contact.html')

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username == 'admin' and password == 'ADMIN':
            workbook = openpyxl.load_workbook(prediction_data_file)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return render_template('admin_data.html', data=data)
        else:
            flash('Invalid username or password', 'error')
            return render_template('admin.html')
    
    return render_template('admin.html')

@app.route("/predict", methods=['POST'])
def predict():
    if 'username' not in session:
        return redirect(url_for('home'))

    if request.method == 'POST':
        sg = float(request.form['sg'])
        htn = float(request.form['htn'])
        hemo = float(request.form['hemo'])
        dm = float(request.form['dm'])
        al = float(request.form['al'])
        appet = float(request.form['appet'])
        rc = float(request.form['rc'])
        pc = float(request.form['pc'])

        values = np.array([[sg, htn, hemo, dm, al, appet, rc, pc]])
        prediction = model.predict(values)

        workbook = openpyxl.load_workbook(prediction_data_file)
        sheet = workbook.active
        username = session.get('username', 'Unknown')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([username, timestamp, sg, htn, hemo, dm, al, appet, rc, pc, int(prediction[0])])
        workbook.save(prediction_data_file)

        session['user_input'] = {
            'sg': sg, 'htn': htn, 'hemo': hemo, 'dm': dm,
            'al': al, 'appet': appet, 'rc': rc, 'pc': pc
        }
        session['prediction'] = int(prediction[0])

        return render_template('result.html', prediction=prediction)

@app.route('/download_csv')
def download_csv():
    try:
        si = StringIO()
        workbook = openpyxl.load_workbook(prediction_data_file)
        sheet = workbook.active
        
        writer = csv.writer(si)
        headers = ['Username', 'Timestamp', 'Specific Gravity', 'Hypertension', 
                  'Hemoglobin', 'Diabetes Mellitus', 'Albumin', 'Appetite', 
                  'Red Blood Cells', 'Pus Cell', 'Prediction']
        writer.writerow(headers)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            writer.writerow(row)
        
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = "attachment; filename=patient_data.csv"
        output.headers["Content-type"] = "text/csv"
        return output
        
    except Exception as e:
        flash('Error downloading file: ' + str(e))
        return redirect(url_for('admin'))

@app.route("/download_report")
def download_report():
    if 'username' not in session:
        return redirect(url_for('home'))

    user_input = session.get('user_input', {})
    prediction = session.get('prediction')
    username = session.get('username', 'Unknown User')
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Create comparison chart
    chart_buffer = create_comparison_chart(user_input)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        alignment=1,
        spaceAfter=30,
        textColor=HexColor('#2c3e50')
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=HexColor('#34495e'),
        spaceBefore=20,
        spaceAfter=20
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=12,
        textColor=HexColor('#2c3e50'),
        spaceBefore=6,
        spaceAfter=6
    )
    
    prediction_style = ParagraphStyle(
        'Prediction',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.white,
        backColor=HexColor('#27ae60') if prediction == 0 else HexColor('#c0392b'),
        alignment=1,
        spaceBefore=20,
        spaceAfter=20,
        padding=20
    )

    # Content
    content = []
    
    # Header
    content.append(Paragraph("Kidney Health Assessment Report", title_style))
    content.append(Paragraph(f"Generated for: {username}", subtitle_style))
    content.append(Paragraph(f"Date: {timestamp}", normal_style))
    content.append(Spacer(1, 0.25*inch))

    # Test Results Table
    content.append(Paragraph("Your Test Results", subtitle_style))
    
    # Prepare table data with reference ranges
    table_data = [['Parameter', 'Your Value', 'Healthy Range', 'Unit']]
    
    for key, value in user_input.items():
        if key in HEALTHY_RANGES:
            range_info = HEALTHY_RANGES[key]
            table_data.append([
                key.upper(),
                f"{value:.2f}",
                f"{range_info['min']} - {range_info['max']}",
                range_info['unit']
            ])
    
    table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#ecf0f1'), colors.white]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    content.append(table)
    
    # Add comparison chart
    content.append(Paragraph("Value Comparison Chart", subtitle_style))
    img = Image(chart_buffer, width=6*inch, height=4*inch)
    content.append(img)
    
    # Prediction Result
    content.append(Spacer(1, 0.25*inch))
    prediction_text = ('Based on our analysis, your results indicate a lower risk of Chronic Kidney Disease. ' 
                      ) if prediction == 0 else (
                      'Based on our analysis, you may have indicators of Chronic Kidney Disease. '
                      )
    content.append(Paragraph(prediction_text, prediction_style))
    
    # Disclaimer
    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=HexColor('#7f8c8d'),
        alignment=1,
        spaceBefore=30
    )
    content.append(Paragraph(
        "Disclaimer: This report is generated by an AI model and should not be considered as a medical diagnosis. "
        "Always consult with qualified healthcare professionals for proper medical advice and treatment.",
        disclaimer_style
    ))

    # Build PDF
    doc.build(content)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="kidney_health_report.pdf",
        mimetype="application/pdf"
    )
@app.route('/logout')
def logout():
    session.clear()  # Clear all session data
    response = make_response(redirect(url_for('home')))
    return response

if __name__ == "__main__":
    app.run(debug=True)